import carddraw
from carddraw import *
import commands
from openpyxl import load_workbook

filename = "Blackjack.xlsx"
workbook = load_workbook(filename)


# Lists all save files
print("What file would you like to access?")
print("For a new save file, please type a new name")
for sheet_name in workbook.sheetnames[1:]:
    print(sheet_name)

# Asks the user which file they would like to access
sheet_name = input("> ")

# Check if sheet exists
if sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]
else:
    # If the sheet does not exist, create a new sheet with that name
    source_sheet = workbook["Source Sheet"]
    worksheet = workbook.create_sheet(sheet_name)

    for row in source_sheet.iter_rows(values_only=True):
        worksheet.append(row)

    workbook.save(filename)

success_cell = worksheet["E3"]
failure_cell = worksheet["E4"]
current_balance = worksheet["B3"]

startup = True
running = False
gambling_money = current_balance.value

while True:
    while startup:
        carddraw.player_cards = []
        carddraw.dealers_cards = []
        carddraw.drawn_cards = []

        print()
        print(f"You have ${gambling_money}")
        # Asks the user if they want to play
        start = input("Would you like to play a round? Y/N > ")
        if gambling_money < 6:
            print("You're out of money!")
            sheet_to_delete = workbook[sheet_name]
            workbook.remove(sheet_to_delete)
            workbook.save(filename)
            print(f"File '{sheet_name}' has been deleted due to lack of funds.")
            quit()

        if start.lower() == "y":
            ante_met = False

            # Ante

            while not ante_met:

                try:
                    ante = int(input("How much would you like to ante? (minimum $6) > "))
                except ValueError:
                    continue
                if ante < 6 or ante > gambling_money:
                    print(f"Please input a value between 6 and {gambling_money}")
                else:
                    break
            gambling_money -= ante

            print(f"Okay, you have ${gambling_money} left.")

            running = True

            # Draws players cards
            draw_player_card()

            draw_dealer_card()

            draw_player_card()

            print(f"Your total is {(commands.player_total())}")

            draw_dealer_card()

            print(f"The dealer has a {carddraw.dealers_cards[~1]}")

            # Exits starting commands
            startup = False

            # Blackjack
            if commands.real_dealer_total(d_total=0) == 21 and not commands.player_total() == 21:
                print("House hit Blackjack :( ")
                print(carddraw.dealers_cards)
                failure_current_value = failure_cell.value
                failure_new_value = failure_current_value + 1
                failure_cell.value = failure_new_value
                current_balance.value = gambling_money
                workbook.save(filename)
                running = False
                startup = True

            if commands.player_total() == 21:
                print("Blackjack!")
                running = False
                startup = True
                gambling_money += ante * 5 / 2
                print(f"Congrats! You win ${ante * 5 / 2}")
                success_current_value = success_cell.value
                success_new_value = success_current_value + 1
                success_cell.value = success_new_value
                current_balance.value = gambling_money
                workbook.save(filename)

        # Exits program if player does not want to play
        elif start.lower() == "n":
            startup = False
            print("Okay. Have a good day")
            quit()
        else:
            print("Please input Y or N")

    # If user responds yes to play, and no BlackJack occurs, run the game
    while running:
        print("Type 'help' to get started!")
        command = input(">")

        # Command List

        if command.lower() == "help":
            print(f"Available Commands: {commands.command_list}")

        elif command.lower() == "hit":
            draw_player_card()
            if commands.player_total() > 21:
                print("Bust! You lose.")
                print(f"You have ${gambling_money} remaining.")
                failure_current_value = failure_cell.value
                failure_new_value = failure_current_value + 1
                failure_cell.value = failure_new_value
                current_balance.value = gambling_money
                workbook.save(filename)

                startup = True
                running = False

            elif commands.player_total() == 21:
                print("Nice! you win!")
                gambling_money += ante * 2
                print(f"You new balance is: ${gambling_money}")
                success_current_value = success_cell.value
                success_new_value = success_current_value + 1
                success_cell.value = success_new_value
                current_balance.value = gambling_money
                workbook.save(filename)

                startup = True
                running = False
            print(f"You have a total of {commands.player_total()}")

        elif command.lower() == "show hand":
            commands.show_hand()

        elif command.lower() == "show dealers hand":
            commands.show_dealer_hand()

        elif command.lower() == "instructions":
            commands.instructions()

        elif command.lower() == "my total":
            total = commands.player_total()
            print(total)

        elif command.lower() == "dealers total":
            print(commands.print_dealer_total(d_total=0))

        elif command.lower() == "stay":
            dealer_below = False

            # If dealer has less than 17, they must hit.
            if commands.real_dealer_total(d_total=0) < 17:
                dealer_below = True

            while dealer_below:
                draw_dealer_card()
                print(f"The house draws a {carddraw.dealers_cards[~1]}")
                if commands.real_dealer_total(d_total=0) >= 17:
                    dealer_below = False

            if commands.real_dealer_total(d_total=0) == 21:
                print("House hit 21.")
                print(carddraw.dealers_cards)

                running = False
                startup = True
                if commands.real_dealer_total(d_total=0) and not commands.player_total():
                    print(f"You lose. You have ${gambling_money} remaining.")

            elif 17 <= commands.real_dealer_total(d_total=0) < 21:
                print()
                print("The house stays")
                print(carddraw.dealers_cards)
                print(f"The house has {commands.real_dealer_total(d_total=0)}")
                print()
                if commands.real_dealer_total(d_total=0) < commands.player_total():
                    print(f"Congrats! You win. You had {commands.player_total()} and the house had {commands.real_dealer_total(d_total=0)}")
                    gambling_money += ante * 2
                    success_current_value = success_cell.value
                    success_new_value = success_current_value + 1
                    success_cell.value = success_new_value
                    current_balance.value = gambling_money
                    workbook.save(filename)
                if commands.real_dealer_total(d_total=0) > commands.player_total():
                    print(f"You lose. You had {commands.player_total()} and the house had {commands.real_dealer_total(d_total=0)}")
                    failure_current_value = failure_cell.value
                    failure_new_value = failure_current_value + 1
                    failure_cell.value = failure_new_value
                    current_balance.value = gambling_money
                    workbook.save(filename)
                if commands.real_dealer_total(d_total=0) == commands.player_total():
                    print(
                        f"Draw. You had {commands.player_total()} and the house had {commands.real_dealer_total(d_total=0)}")
                    gambling_money += ante
                running = False
                startup = True

            elif commands.real_dealer_total(d_total=0) > 21:
                print()
                print("Congrats you win! The house busts.")
                success_current_value = success_cell.value
                success_new_value = success_current_value + 1
                success_cell.value = success_new_value

                print(carddraw.dealers_cards)
                print(f"The House had a total of {commands.real_dealer_total(d_total=0)}")
                gambling_money += ante * 2
                print(f"You now have a total of ${gambling_money}")
                running = False
                startup = True
                current_balance.value = gambling_money
                workbook.save(filename)
        else:
            print("Please put in a command")


